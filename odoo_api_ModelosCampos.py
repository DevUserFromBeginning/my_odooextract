import xmlrpc.client
import openpyxl
import pprint


'''
Realizado por: Ernesto Caraballo
Fecha actual: 10/05/2024

.- Conectarme a la api (OK)
.- Modelos disponibles: Faltan los modelos de manejo inventario...I'll make it anyway
.- Campos de cada modelo (OK, de los que tengo hasta ahora 10/02/2024)
.- Campos útiles de cada modelo para cada caso (852, GPOS, etc)
.- Registros de cada modelo
.- Cómo se relacionan los modelos?
.- Llevar los registros a excel
.- LLevar la data a una bd desnormalizada
.- Actualización de la data
.- Cómo registra ODOO las actualizaciones?

'''

def main():
    url = 'https://importvzla-import-import-prueba-12847814.dev.odoo.com'
    db = 'importvzla-import-import-prueba-12847814'
    username = 'admin'
    password = 'Import2023!'

    common = xmlrpc.client.ServerProxy('{}/xmlrpc/2/common'.format(url))
    version = common.version()

    uid = common.authenticate(db, username, password, {})

    models = xmlrpc.client.ServerProxy('{}/xmlrpc/2/object'.format(url))

    available_models = ['crm.lead', 'res.partner', 'res.users', 'sale.order', 'sale.order.line', 'product.template', 'product.product', 
                        'purchase.order', 'purchase.order.line', 'x_linea', 'x_subninea', 'x_fabricante','stock.move','stock.move.line',
                        'stock.picking.type','stock.picking','stock.warehouse.orderpoint','stock.quant']
    # stock.picking transferencias de inventario
    # stock.warehouse.orderpoint reposiciones de inventario
    # stock.quant ajustes de inventario
    
 
    wb = openpyxl.Workbook()
    
    for modelo in available_models:
        # get each model fields into a list
        wb.create_sheet(modelo)
        campos = models.execute_kw(db, uid, password, modelo, 'fields_get',[],{'attributes': ['string']})
        columna = 1
        fila = 1
        ws = wb[modelo]
        for campo in campos.keys():
            ws.cell(row=fila, column=columna).value = campo
            fila= fila + 1
            
    wb.save('Modelos_Campos.xlsx')
    

if __name__ == '__main__':
    main()