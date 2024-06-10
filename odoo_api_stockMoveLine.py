import xmlrpc.client
import openpyxl


'''
Realizado por: Ernesto Caraballo
Fecha actual: 10/05/2024

.- Conectarme a la api (OK)
.- Extraer los registros del modelo: crm.leads
.- Determinas los campos útiles del modelo para cada caso (852, GPOS, etc)
.- Llevar los registros a excel

    username = 'admin'
    password = 'Import2023!'
'''

def main():
    url = 'https://importvzla-import-import-prueba-12847814.dev.odoo.com'
    db = 'importvzla-import-import-prueba-12847814'
    username = 'francisco.tellez@import-import.com'
    password = '1129734'

    common = xmlrpc.client.ServerProxy('{}/xmlrpc/2/common'.format(url))
    version = common.version()

    uid = common.authenticate(db, username, password, {})

    models = xmlrpc.client.ServerProxy('{}/xmlrpc/2/object'.format(url))
    
    ####
    # en esta consulta, por fin logré, utilizar "o" en la condicion
    ####
    # me falta condicionar la fecha de la data; no logro que funcione el filtro por fecha
    ordersLines = models.execute_kw(db, uid, password,'stock.move.line', 'search_read',
                                 [[['picking_code','not in',['internal',False]],['qty_done','!=',0],['state','!=','cancel'],
                                   '|',('picking_type_id','like','Principal: Expediciones'),('picking_type_id','like', 'Operador Logistico: Recepciones')]],
                                 {'fields':['id','move_id','origin','__last_update','display_name','product_id','location_id','location_dest_id',
                                            'picking_type_id','picking_partner_id','qty_done','product_uom_id','sale_price',
                                            'picking_id','state']})
                            
    '''
    contador = 1
    for orderLine in ordersLines:
        print(f'item: {contador}')
        print(orderLine)
        print('\n')
        contador +=1
    '''
    
    campos = []
    for campo in ordersLines[0].keys():
        campos.append(campo)
        
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "stock.move.line"
    
    
    tmpRow = 1
    tmpColumn = 1
    for _ in range(0, len(campos)):
        ws.cell(row=1, column=_+1).value = campos[_]
        
    tmpRow = tmpRow + 1 #to make content start filling at row 2
    tmpColumn = 1
    
    for order in ordersLines:
        for campo in campos:
            ws.cell(row=tmpRow, column=tmpColumn).value= str(order[campo])
            tmpColumn = tmpColumn + 1
        tmpColumn=1
        tmpRow = tmpRow + 1
    
    wb.save(r"C:\\Users\\ESCH\Desktop\\odoo-852\\excel\\stockMoveLine.xlsx")        
    
    print(f"\n********************\nSe acab´lo que se daba") 
    '''
    
if __name__ == '__main__':
    main()
    
