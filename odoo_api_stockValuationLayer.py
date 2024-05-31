import xmlrpc.client
import openpyxl


'''
Realizado por: Ernesto Caraballo
Fecha actual: 10/05/2024

.- Conectarme a la api (OK)
.- Extraer los registros del modelo: crm.leads
.- Determinas los campos útiles del modelo para cada caso (852, GPOS, etc)
.- Llevar los registros a excel
    
    username = 'admin'
    password = 'Import2023!'
'''

def main():
    url = 'https://importvzla-import-import-prueba-12847814.dev.odoo.com'
    db = 'importvzla-import-import-prueba-12847814'
    username = 'francisco.tellez@import-import.com'
    password = '1129734'

    common = xmlrpc.client.ServerProxy('{}/xmlrpc/2/common'.format(url))
    version = common.version()

    uid = common.authenticate(db, username, password, {})

    models = xmlrpc.client.ServerProxy('{}/xmlrpc/2/object'.format(url))
    
    #['company_id','=','Import Import'],['state','in',['waiting','confirmed','assigned']],['location_dest_id','=','Partner Locations/Customers']
    ordersLines = models.execute_kw(db, uid, password,'stock.valuation.layer', 'search_read',
                                 [[['product_id','ilike','Hamburguesa de queso RefInt']]],
                                 {'fields':['id','product_id','quantity','uom_id','currency_id','unit_cost','value','remaining_value','description',
                                            'stock_move_id','__last_update','display_name','x_studio_documento_de_origen']})
    
    '''
    contador = 0
    for orderLine in ordersLines:
        print(f'item {contador}: ')
        print(orderLine)
        print('\n')
        contador +=1
    '''
    
    campos = []
    for campo in ordersLines[0].keys():
        campos.append(campo) 
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title='stock.valuation.layer'
    
    tmpRow = 1
    tmpColumn = 1
    for _ in range(0, len(campos)):
        ws.cell(row=1, column=_+1).value = campos[_]
        
    tmpRow = tmpRow + 1 #to make content start filling at row 2
    tmpColumn = 1
    
    for order in ordersLines:
        for campo in campos:
            ws.cell(row=tmpRow, column=tmpColumn).value= str(order[campo])
            tmpColumn = tmpColumn + 1
        tmpColumn=1
        tmpRow = tmpRow + 1
    
    wb.save(r"C:\\Users\\ESCH\Desktop\\odoo-852\\excel\\stockValuationLayer.xlsx")        
    
    print(f"\n********************\nSe acab´lo que se daba")
    
    
if __name__ == '__main__':
    main()