import xmlrpc.client
import openpyxl

'''
Realizado por: Ernesto Caraballo
Fecha actual: 16/05/2024

.- Conectarme a la api (OK)
.- Extraer los registros del modelo: crm.leads
.- Determinas los campos útiles del modelo para cada caso (852, GPOS, etc)
.- Llevar los registros a excel

'''

def main():
    url = 'https://importvzla-import-import-prueba-12847814.dev.odoo.com'
    db = 'importvzla-import-import-prueba-12847814'
    username = 'admin'
    password = 'Import2023!'

    common = xmlrpc.client.ServerProxy('{}/xmlrpc/2/common'.format(url))
    version = common.version()

    uid = common.authenticate(db, username, password, {})

    models = xmlrpc.client.ServerProxy('{}/xmlrpc/2/object'.format(url))

    leads = models.execute_kw(db, uid, password,'crm.lead', 'search_count',[[['company_id', '=', 'Import Import']]])
    print(leads)
    
    
    leads = models.execute_kw(db, uid, password,'crm.lead', 'search_read',
                                 [[['company_id', '=', 'Import Import'],['id','<','200'],['name','like','TOPIECA']]],
                                 {'fields': ['id','company_id','create_date','__last_update','name','description','display_name','type','stage_id','date_last_stage_update',
                                             'partner_id','partner_name','contact_name','email_from','probability','expected_revenue','company_currency','user_id',
                                             'date_deadline','date_closed','lost_reason','won_status', 'days_exceeding_closing', 'sale_amount_total', 'quotation_count', 'sale_order_count', 
                                             'order_ids', 'x_linea','x_origen','x_industria','x_petroleo','x_grupo','x_zona','x_canal'],'limit':5})
    
    
    
    for partner in leads:
        print(partner)
        print('\n')
    
    '''
    campos = []
    for campo in leads[0].keys():
        campos.append(campo)
        
    wb = openpyxl.Workbook()
    wb.create_sheet("crm.lead")
    ws = wb['crm.lead']
    
    tmpRow = 1
    tmpColumn = 1
    for _ in range(0, len(campos)):
        ws.cell(row=1, column=_+1).value = campos[_]
        
    tmpRow = tmpRow + 1 #to make content start filling at row 2
    tmpColumn = 1
    
    for order in leads:
        for campo in campos:
            ws.cell(row=tmpRow, column=tmpColumn).value= str(order[campo])
            tmpColumn = tmpColumn + 1
        tmpColumn=1
        tmpRow = tmpRow + 1
    
    wb.save(r"C:\\Users\\ESCH\Desktop\\odoo\\excel\\crmLead.xlsx")        
    
    print(f"\n********************\nSe acab´lo que se daba")
    '''
     
if __name__ == '__main__':
    main()
    
    
