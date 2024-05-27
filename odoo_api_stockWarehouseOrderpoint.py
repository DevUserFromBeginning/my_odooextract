import xmlrpc.client
import openpyxl
import pprint


'''
Realizado por: Ernesto Caraballo
Fecha actual: 10/05/2024

.- Conectarme a la api (OK)
.- Extraer los registros del modelo: crm.leads
.- Determinas los campos útiles del modelo para cada caso (852, GPOS, etc)
.- Llevar los registros a excel

'''

def main():
    url = 'https://importvzla-import-import-prueba-12847814.dev.odoo.com'
    db = 'importvzla-import-import-prueba-12847814'
    username = 'admin'
    password = 'Import2023!'

    common = xmlrpc.client.ServerProxy('{}/xmlrpc/2/common'.format(url))
    version = common.version()

    uid = common.authenticate(db, username, password, {})

    models = xmlrpc.client.ServerProxy('{}/xmlrpc/2/object'.format(url))
    
    #['company_id', '=', 'Import Import'] # filtro de compañía...aparentemente no funciona
    ordersLines = models.execute_kw(db, uid, password,'stock.warehouse.orderpoint', 'search_read',
                                 [[['company_id','=','Import Import']]],
                                 {'fields':['name','trigger','active','snoozed_until','warehouse_id','location_id','product_tmpl_id','product_id','product_category_id','product_uom',
                                            'product_uom_name','product_min_qty','product_max_qty','qty_multiple','group_id','company_id','allowed_location_ids','rule_ids',
                                            'lead_days_date','route_id','qty_on_hand','qty_forecast','qty_to_order','id','__last_update','display_name','create_uid','create_date',
                                            'write_uid','write_date','show_bom','bom_id','show_supplier','supplier_id','vendor_id'],'limit':400})
    '''for orderLine in ordersLines:
        print(orderLine)
        print('\n')
    '''
    
    campos = []
    for campo in ordersLines[0].keys():
        campos.append(campo)
        
    
    wb = openpyxl.Workbook()
    wb.create_sheet("stock.warehouse.orderpoint")
    ws = wb['stock.warehouse.orderpoint']
    
    tmpRow = 1
    tmpColumn = 1
    for _ in range(0, len(campos)):
        ws.cell(row=1, column=_+1).value = campos[_]
        
    tmpRow = tmpRow + 1 #to make content start filling at row 2
    tmpColumn = 1
    
    for order in ordersLines:
        for campo in campos:
            ws.cell(row=tmpRow, column=tmpColumn).value= str(order[campo])
            tmpColumn = tmpColumn + 1
        tmpColumn=1
        tmpRow = tmpRow + 1
    
    wb.save(r"C:\\Users\\ESCH\Desktop\\odoo_api\\excel\\stockWarehouseOrderpoint.xlsx")        
    
    print(f"\n********************\nSe acab´lo que se daba") 
    
if __name__ == '__main__':
    main()