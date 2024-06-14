import xmlrpc.client
import openpyxl


'''
Realizado por: Ernesto Caraballo
Fecha actual: 05/06/2024

.- Conectarme a la api (OK)
.- Extraer los registros del modelo: account.move
.- Determinas los campos útiles del modelo para cada caso (852, GPOS, etc)
.- Llevar los registros a excel

'''

def main():
    url = 'https://importvzla-import-import-prueba-12847814.dev.odoo.com'
    db = 'importvzla-import-import-prueba-12847814'
    username = 'francisco.tellez@import-import.com'
    password = '1129734'

    common = xmlrpc.client.ServerProxy('{}/xmlrpc/2/common'.format(url))
    version = common.version()

    uid = common.authenticate(db, username, password, {})

    models = xmlrpc.client.ServerProxy('{}/xmlrpc/2/object'.format(url))
    
    total_records = models.execute_kw(db, uid, password, 'account.move.line', 'search_count', [[]])
    
    print(f'total registros: {total_records}')

    remaining_records = total_records
    start = 1 #start no cambia es un ciclo
    block = 10000 #block cambia sólo con el último grupo de registros

    rcontador = 0
    while True:
        if remaining_records == 0:
            break
        elif block > remaining_records:
            block = remaining_records
        else:
            ordersLines = models.execute_kw(db, uid, password,'account.move.line', 'search_read',
                                    [[]],
                                    {'fields':['id','date','create_date','write_date','__last_update','move_id','move_name','ref','journal_id','account_id','account_internal_type',
                                               'account_internal_group','name','quantity','price_unit','debit','credit','balance','amount_currency','price_subtotal','price_total',
                                               'date_maturity','currency_id','partner_id','product_id','payment_id','tax_ids','tax_line_id','tax_base_amount','amount_residual',
                                               'amount_residual_currency','matched_debit_ids','matched_credit_ids','matching_number','purchase_line_id','purchase_order_id',
                                               'sale_line_ids','product_type','x_studio_related_field_4zjYm','x_studio_fabricante'],
                                     'offset':rcontador, 'limit':block})
                
            #pasar a la hoja de excel
            
            if rcontador == 0:
            
                # lista con los nombres de los campos
                campos = []
                for campo in ordersLines[0].keys():
                    campos.append(campo)

                # crear el libro
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = 'account.moveLine'

                #grabar los campos en la primera fila de la hoja activa
                tmpRow = 1
                tmpColumn = 1
                for _ in range(0, len(campos)):
                    ws.cell(row=1, column=_+1).value = campos[_]
                    
                
            # vaciar el contenido de ordersLines a la hoja                
            tmpRow = tmpRow + 1 #to make content start filling at row 2
            tmpColumn = 1
        
            for order in ordersLines:
                for campo in campos:
                    ws.cell(row=tmpRow, column=tmpColumn).value= str(order[campo])
                    tmpColumn = tmpColumn + 1
                tmpColumn=1
                tmpRow = tmpRow + 1
            
            
            #actualizar estado de variables    
            remaining_records = remaining_records - block
            rcontador = rcontador + block
            print(f'start: {start} remaining_records: {remaining_records} rcontador: {rcontador} block: {block}\n')
            
        wb.save(r"C:\\Users\\ESCH\Desktop\\odoo-852\\excel\\accountMoveLine.xlsx")
    
    print(f"\n********************\nSe acab´lo que se daba")
    
    
if __name__ == '__main__':
    main()
    
    

