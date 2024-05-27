import xmlrpc.client
import openpyxl


'''
Realizado por: Ernesto Caraballo
Fecha actual: 10/05/2024

.- Conectarme a la api (OK)
.- Extraer los registros del modelo: sale.order.line
.- Determinas los campos útiles del modelo para cada caso (852, GPOS, etc)
.- Llevar los registros a excel

'''

def main():
    url = 'https://importvzla-import-import-prueba-12847814.dev.odoo.com'
    db = 'importvzla-import-import-prueba-12847814'
    username = 'admin'
    password = 'Import2023!'

    common = xmlrpc.client.ServerProxy('{}/xmlrpc/2/common'.format(url))
    version = common.version()

    uid = common.authenticate(db, username, password, {})

    models = xmlrpc.client.ServerProxy('{}/xmlrpc/2/object'.format(url))
    
    #['company_id', '=', 'Import Import'] # filtro de compañía...aparentemente no funciona
    ordersLines = models.execute_kw(db, uid, password,'sale.order.line', 'search_read',
                                 [[]],
                                 {'fields':['id','company_id','order_id','name','invoice_lines','invoice_status','product_id','product_template_id','price_unit',
                                            'price_subtotal','price_tax','price_total','price_reduce','tax_id','price_reduce_taxinc','price_reduce_taxexcl',
                                            'discount','product_uom_qty','product_uom','qty_delivered','qty_delivered_manual','qty_to_invoice','qty_invoiced',
                                            'untaxed_amount_invoiced','untaxed_amount_to_invoice','salesman_id','currency_id','order_partner_id','is_expense',
                                            'is_downpayment','state','__last_update','display_name','create_uid','create_date', 'sale_order_option_ids',
                                            'purchase_line_ids','purchase_line_count','move_ids','product_type','scheduled_date','forecast_expected_date',
                                            'free_qty_today','qty_available_today','warehouse_id','qty_to_deliver','product_qty','margin','margin_percent',
                                            'purchase_price','x_studio_related_field_AJ5Vx','x_studio_sublinea']})
    
    '''
    for orderLine in ordersLines:
        print(orderLine)
        print('\n')
    '''
    
    campos = []
    for campo in ordersLines[0].keys():
        campos.append(campo)
        
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title='sale.order.line'
    
    tmpRow = 1
    tmpColumn = 1
    for _ in range(0, len(campos)):
        ws.cell(row=1, column=_+1).value = campos[_]
        
    tmpRow = tmpRow + 1 #to make content start filling at row 2
    tmpColumn = 1
    
    for order in ordersLines:
        for campo in campos:
            ws.cell(row=tmpRow, column=tmpColumn).value= str(order[campo])
            tmpColumn = tmpColumn + 1
        tmpColumn=1
        tmpRow = tmpRow + 1
    
    wb.save(r"C:\\Users\\ESCH\Desktop\\odoo\\excel\\saleOrderLine.xlsx")        
    
    print(f"\n********************\nSe acab´lo que se daba")
    
    
if __name__ == '__main__':
    main()