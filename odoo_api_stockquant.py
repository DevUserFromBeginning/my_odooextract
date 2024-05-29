import xmlrpc.client
import openpyxl


'''
Realizado por: Ernesto Caraballo
Fecha actual: 10/05/2024

.- Conectarme a la api (OK)
.- Extraer los registros del modelo: stock.quant
.- Determinas los campos útiles del modelo para cada caso (852, GPOS, etc)
.- Llevar los registros a excel

'''

def main():
    url = 'https://importvzla-import-import-prueba-12847814.dev.odoo.com'
    db = 'importvzla-import-import-prueba-12847814'
    username = 'francisco.tellez@import-import.com'
    password = '1129734'

    common = xmlrpc.client.ServerProxy('{}/xmlrpc/2/common'.format(url))
    version = common.version()

    uid = common.authenticate(db, username, password, {})

    models = xmlrpc.client.ServerProxy('{}/xmlrpc/2/object'.format(url))
    
    #['company_id', '=', 'Import Import'] # filtro de compañía...aparentemente no funciona
    ordersLines = models.execute_kw(db, uid, password,'stock.quant', 'search_read',
                                 [[['company_id','=','Import Import'],['location_id','in',['INC/Stock','Impor/Stock','OP/Stock','TR/Stock','Impor/Stock/Stock']],['product_id','=like','Hamburguesa de queso']]],
                                 {'fields':['id','company_id','product_id','product_tmpl_id','display_name','product_uom_id','location_id','owner_id',
                                            'quantity','reserved_quantity','available_quantity','in_date','on_hand','product_categ_id','inventory_quantity',
                                            'inventory_diff_quantity','inventory_date','inventory_quantity_set','is_outdated','user_id','__last_update',
                                            'create_uid','create_date','value','currency_id']})
    
    contador = 1
    for orderLine in ordersLines:
        print(f"{contador}.- ")
        print(orderLine)
        print('\n')
        contador +=1
    '''
    
    campos = []
    for campo in ordersLines[0].keys():
        campos.append(campo)
        
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.tittle = "stock.quant"
    
    tmpRow = 1
    tmpColumn = 1
    for _ in range(0, len(campos)):
        ws.cell(row=1, column=_+1).value = campos[_]
        
    tmpRow = tmpRow + 1 #to make content start filling at row 2
    tmpColumn = 1
    
    for order in ordersLines:
        for campo in campos:
            ws.cell(row=tmpRow, column=tmpColumn).value= str(order[campo])
            tmpColumn = tmpColumn + 1
        tmpColumn=1
        tmpRow = tmpRow + 1
    
    wb.save(r"C:\\Users\\ESCH\Desktop\\odoo\\excel\\stockQuant.xlsx")        
    
    print(f"\n********************\nSe acab´lo que se daba")
    '''
    
if __name__ == '__main__':
    main()