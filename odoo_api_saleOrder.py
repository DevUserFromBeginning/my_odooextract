import xmlrpc.client
import openpyxl


'''
Realizado por: Ernesto Caraballo
Fecha actual: 10/05/2024

.- Conectarme a la api (OK)
.- Extraer los registros del modelo: sale.order
.- Determinas los campos útiles del modelo para cada caso (852, GPOS, etc)
.- Llevar los registros a excel

'''

def main():
    url = 'https://importvzla-import-import-prueba-12847814.dev.odoo.com'
    db = 'importvzla-import-import-prueba-12847814'
    username = 'francisco.tellez@import-import.com'
    password = '1129734'

    common = xmlrpc.client.ServerProxy('{}/xmlrpc/2/common'.format(url))
    version = common.version()

    uid = common.authenticate(db, username, password, {})

    models = xmlrpc.client.ServerProxy('{}/xmlrpc/2/object'.format(url))
    

    ordersLines = models.execute_kw(db, uid, password,'sale.order', 'search_read',
                                 [[['type_name','=','sale_order']]],
                                 {'fields':['id','date_order','name','origin','client_order_ref','reference','state','validity_date','is_expired',
                                            'create_date','user_id','partner_id','pricelist_id','currency_id',
                                            'order_line','invoice_count','invoice_ids','invoice_status','amount_untaxed','amount_tax','amount_total',
                                            'currency_rate','payment_term_id','tax_country_id','commitment_date','expected_date',
                                            'amount_undiscounted','type_name','__last_update',
                                            'display_name','purchase_order_count','warehouse_id','picking_ids','delivery_count',
                                            'procurement_group_id','effective_date','margin','x_npedido','x_studio_n_orden_de_compra',
                                            'x_Fecha','x_totaltn','x_studio_nota_de_entrega','x_studio_fecha_1','x_studio_n_peticin_de_oferta','x_Pedido',
                                            'x_notaentrega','x_studio_total_','x_studio_tipo_de_cotizacin',
                                            'x_studio_factor_de_negociacin','x_ocompra','x_tasav','x_studio_tasa_venta','x_studio_tasa_ref_negociada',
                                            'x_fechav']}) 
    
    '''
    for orderLine in ordersLines:
        print(orderLine)
        print('\n')
    '''    
    
    campos = []
    for campo in ordersLines[0].keys():
        campos.append(campo)
        
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'sale.order'
    
    tmpRow = 1
    tmpColumn = 1
    for _ in range(0, len(campos)):
        ws.cell(row=1, column=_+1).value = campos[_]
        
    tmpRow = tmpRow + 1 #to make content start filling at row 2
    tmpColumn = 1
    
    for order in ordersLines:
        for campo in campos:
            ws.cell(row=tmpRow, column=tmpColumn).value= str(order[campo])
            tmpColumn = tmpColumn + 1
        tmpColumn=1
        tmpRow = tmpRow + 1
    
    wb.save(r"C:\\Users\\ESCH\Desktop\\odoo-852\\excel\\saleOrder.xlsx")
            
    
    print(f"\n********************\nSe acab´lo que se daba")
    
    
if __name__ == '__main__':
    main()