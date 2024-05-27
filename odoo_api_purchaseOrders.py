import xmlrpc.client
import openpyxl


'''
Realizado por: Ernesto Caraballo
Fecha actual: 10/05/2024

.- Conectarme a la api (OK)
.- Extraer los registros del modelo: crm.leads
.- Determinas los campos útiles del modelo para cada caso (852, GPOS, etc)
.- Llevar los registros a excel

'''

def main():
    url = 'https://importvzla-import-import-prueba-12847814.dev.odoo.com'
    db = 'importvzla-import-import-prueba-12847814'
    username = 'admin'
    password = 'Import2023!'

    common = xmlrpc.client.ServerProxy('{}/xmlrpc/2/common'.format(url))
    version = common.version()

    uid = common.authenticate(db, username, password, {})

    models = xmlrpc.client.ServerProxy('{}/xmlrpc/2/object'.format(url))
 
    ordersLines = models.execute_kw(db, uid, password,'purchase.order', 'search_read',
                                 [[['company_id', '=', 'Import Import']]],
                                 {'fields': ['id','company_id','name','date_order','date_approve','partner_id','currency_id','state','order_line','notes','invoice_count',
                                             'invoice_ids','invoice_status','date_planned','amount_untaxed','amount_tax',
                                             'amount_total','tax_country_id','payment_term_id','product_id','user_id','currency_rate','__last_update',
                                             'display_name','create_uid','create_date','write_uid','write_date','requisition_id','incoming_picking_count','picking_ids',
                                             'sale_order_count','purchase_manual_currency_rate_active','x_studio_fecha_1','x_studio_fecha','x_studio_fecha_2',
                                             'x_studio_total__1','x_OC','x_studio_tasa_compra','x_fechac','x_Factura','x_studio_nota_de_recepcion']})
    '''
    for orderLine in ordersLines:
        print(orderLine)
        print('\n')
    '''
    campos = []
    for campo in ordersLines[0].keys():
        campos.append(campo)
        
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'purchase.order'
    
    tmpRow = 1
    tmpColumn = 1
    for _ in range(0, len(campos)):
        ws.cell(row=1, column=_+1).value = campos[_]
        
    tmpRow = tmpRow + 1 #to make content start filling at row 2
    tmpColumn = 1
    
    for order in ordersLines:
        for campo in campos:
            ws.cell(row=tmpRow, column=tmpColumn).value= str(order[campo])
            tmpColumn = tmpColumn + 1
        tmpColumn=1
        tmpRow = tmpRow + 1
    
    wb.save(r"C:\\Users\\ESCH\Desktop\\odoo\\excel\\purchaseOrder.xlsx")        
    
    print(f"\n********************\nSe acab´lo que se daba")
    
    
if __name__ == '__main__':
    main()
    

