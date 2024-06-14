'''
Realizado por: Ernesto Caraballo
Fecha actual: 05/06/2024

.- Conectarme a la api (OK)
.- Extraer los registros del modelo: account.move
.- Determinas los campos útiles del modelo para cada caso (852, GPOS, etc)
.- Llevar los registros a excel

'''
import xmlrpc.client
import openpyxl


def load_from_API():
    '''
    Función para consultar y traer todos los artículos del modelo product.template
    
    :param: no recibe parámetros
    :return: devuelve una lista cuyos elementos son diccionarios con todos los registros devueltos por la API
    '''
    url = 'https://importvzla-import-import-prueba-12847814.dev.odoo.com'
    db = 'importvzla-import-import-prueba-12847814'
    username = 'francisco.tellez@import-import.com'
    password = '1129734'

    common = xmlrpc.client.ServerProxy('{}/xmlrpc/2/common'.format(url))
    version = common.version()

    uid = common.authenticate(db, username, password, {})

    models = xmlrpc.client.ServerProxy('{}/xmlrpc/2/object'.format(url))
    

    #ordersLines = models.execute_kw(db, uid, password, 'account.move', 'fields_get',[],{'attributes': ['string']})
    
    # this is filtered for out_invoices, sales **** ('move_type','=','out_invoice'), ****
    ordersLines = models.execute_kw(db, uid, password,'account.move', 'search_read',
                                    [[]],
                                    {'fields':['id','sequence_number','name','highest_name','date','ref','journal_id',
                                               'suitable_journal_ids','line_ids','partner_id','move_type','country_code','payment_reference',
                                               'amount_untaxed','amount_tax','amount_total', 'amount_residual','payment_state',
                                               'reversal_move_id','invoice_date','invoice_date_due','invoice_origin',
                                               'invoice_payment_term_id','invoice_line_ids','invoice_partner_display_name','__last_update',
                                               'display_name','create_date','attachment_ids','partner_shipping_id','manual_currency_rate',
                                               'fiscal_provider','x_tasa','x_payment_ids','sale_order_id','sale_order_number','rate',
                                               'amount_untaxed_rate','amount_tax_rate','amount_total_rate','x_studio_total_','x_studio_deuda',
                                               'x_tipodoc','x_ncontrol','x_studio_related_field_35Zlc', 'x_studio_orden_de_compra'],'offset':20500,'limit':60})
    return ordersLines
    
def main():
     
    
    '''
    contador = 1
    for orderLine in ordersLines:
        print(f"contador {contador} Id:{orderLine['id']}")
        print('\n************\n')
        contador +=1
    '''  
    
    lines = load_from_API()
    campos = []
    for campo in lines[0].keys():
        campos.append(campo)
        
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'account.move'
    
    tmpRow = 1
    tmpColumn = 1
    for _ in range(0, len(campos)):
        ws.cell(row=1, column=_+1).value = campos[_]
        
    tmpRow = tmpRow + 1 #to make content start filling at row 2
    tmpColumn = 1
    
    for order in lines:
        for campo in campos:
            ws.cell(row=tmpRow, column=tmpColumn).value= str(order[campo])
            tmpColumn = tmpColumn + 1
        tmpColumn=1
        tmpRow = tmpRow + 1
    
    wb.save(r"C:\\Users\\ESCH\Desktop\\odoo-852\\excel\\accountMove.xlsx")
            
    
    print(f"\n********************\nSe acab´lo que se daba")
    
    
if __name__ == '__main__':
    main()
    
    

