import xmlrpc.client
import openpyxl


'''
Realizado por: Ernesto Caraballo
Fecha actual: 16/05/2024

.- Conectarme a la api (OK)
.- Extraer los registros del modelo: purchase.order.line
.- Determinas los campos útiles del modelo para cada caso (852, GPOS, etc)
.- Llevar los registros a excel

'''

def main():
    url = 'https://importvzla-import-import-prueba-12847814.dev.odoo.com'
    db = 'importvzla-import-import-prueba-12847814'
    username = 'admin'
    password = 'Import2023!'

    common = xmlrpc.client.ServerProxy('{}/xmlrpc/2/common'.format(url))
    version = common.version()

    uid = common.authenticate(db, username, password, {})

    models = xmlrpc.client.ServerProxy('{}/xmlrpc/2/object'.format(url))
 
    ordersLines = models.execute_kw(db, uid, password,'purchase.order.line', 'search_read',
                                 [[['company_id', '=', 'Import Import']]],
                                 {'fields': ['id','company_id','order_id','date_order','partner_id','product_type','product_id','name','product_qty','product_uom_qty','date_planned',
                                             'product_uom','product_uom_category_id','price_unit','price_subtotal','price_total','price_tax','state','invoice_lines','qty_invoiced',
                                             'qty_received','qty_received_manual','qty_to_invoice','currency_id','display_name','create_uid','create_date','__last_update']})
    '''
    for orderLine in ordersLines:
        print(orderLine)
        print('\n')
    '''
    campos = []
    for campo in ordersLines[0].keys():
        campos.append(campo)
        
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'purchase.order.line'
    
    tmpRow = 1
    tmpColumn = 1
    for _ in range(0, len(campos)):
        ws.cell(row=1, column=_+1).value = campos[_]
        
    tmpRow = tmpRow + 1 #to make content start filling at row 2
    tmpColumn = 1
    
    for order in ordersLines:
        for campo in campos:
            ws.cell(row=tmpRow, column=tmpColumn).value= str(order[campo])
            tmpColumn = tmpColumn + 1
        tmpColumn=1
        tmpRow = tmpRow + 1
    
    wb.save(r"C:\\Users\\ESCH\Desktop\\odoo\\excel\\purchaseOrderLine.xlsx")        
    
    print(f"\n********************\nSe acab´lo que se daba")
    
if __name__ == '__main__':
    main()
    

