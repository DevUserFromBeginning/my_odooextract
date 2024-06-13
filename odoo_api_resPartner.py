import xmlrpc.client
import openpyxl


'''
Realizado por: Ernesto Caraballo
Fecha actual: 10/05/2024

.- Conectarme a la api (OK)
.- Extraer los registros del modelo: crm.leads
.- Determinas los campos útiles del modelo para cada caso (852, GPOS, etc)
.- Llevar los registros a excel

'''

def main():
    url = 'https://importvzla-import-import.odoo.com'
    db = 'importvzla-import-import-import-5808788'
    username = 'francisco.tellez@import-import.com'
    password = '1129734'

    common = xmlrpc.client.ServerProxy('{}/xmlrpc/2/common'.format(url))
    version = common.version()

    uid = common.authenticate(db, username, password, {})

    models = xmlrpc.client.ServerProxy('{}/xmlrpc/2/object'.format(url))
    
    #['company_id', '=', 'Import Import'] # filtro de compañía...aparentemente no funciona
    # pendiente de este tema 
    ordersLines = models.execute_kw(db, uid, password,'res.partner', 'search_read',
                                 [[]],
                                 {'fields':['id','company_id','name','active','type','street','street2','zip','city',
                                            'state_id','country_id','country_code','is_company','industry_id','company_type',
                                            'contact_address','commercial_partner_id','commercial_company_name','self','__last_update',
                                            'contact_address_complete','credit','debit','debit_limit','total_invoiced','currency_id',
                                            'sale_order_count','unpaid_invoices','total_due','total_overdue',
                                            'x_studio_canal_comercial','x_tipopersona','x_studio_bpid','x_studio_tipo_de_proveedor','x_studio_zonas_de_ventas',
                                            'x_studio_tipo_de_cliente','x_studio_cliente']})
    '''
    for orderLine in ordersLines:
        print(orderLine)
        print('\n')
    '''
    campos = []
    for campo in ordersLines[0].keys():
        campos.append(campo)
        
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'res.partner'
    
    tmpRow = 1
    tmpColumn = 1
    for _ in range(0, len(campos)):
        ws.cell(row=1, column=_+1).value = campos[_]
        
    tmpRow = tmpRow + 1 #to make content start filling at row 2
    tmpColumn = 1
    
    for order in ordersLines:
        for campo in campos:
            ws.cell(row=tmpRow, column=tmpColumn).value= str(order[campo])
            tmpColumn = tmpColumn + 1
        tmpColumn=1
        tmpRow = tmpRow + 1
    
    wb.save(r"C:\\Users\\ESCH\Desktop\\odoo-852\\excel\\resPartner.xlsx")        
    
    print(f"\n********************\nSe acab´lo que se daba")
    
    
if __name__ == '__main__':
    main()