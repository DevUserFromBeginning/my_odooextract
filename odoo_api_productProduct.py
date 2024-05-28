import xmlrpc.client
import openpyxl

'''
Realizado por: Ernesto Caraballo
Fecha actual: 10/05/2024

.- Conectarme a la api (OK)
.- Extraer los registros del modelo: product.product
.- Determinas los campos útiles del modelo para cada caso (852, GPOS, etc)
.- Llevar los registros a excel

'''
# instancia operativa
# https://importvzla-import-import.odoo.com
# importvzla-import-import-import-5808788

# instancia prueba
# 'https://importvzla-import-import-prueba-12847814.dev.odoo.com'
# 'importvzla-import-import-prueba-12847814'
def main():
    url = 'https://importvzla-import-import-prueba-12847814.dev.odoo.com'
    db = 'importvzla-import-import-prueba-12847814'
    username = 'admin'
    password = 'Import2023!'

    common = xmlrpc.client.ServerProxy('{}/xmlrpc/2/common'.format(url))
    version = common.version()

    uid = common.authenticate(db, username, password, {})

    models = xmlrpc.client.ServerProxy('{}/xmlrpc/2/object'.format(url))
    
    # conectar a producción y evaluar los campos significativos para el 852
    # a la fecha (28052024) aún no tengo claro que significa cada uno de los campos cuando 
    # deberían tener significado
    # evaluar los filtros con detenimiento en la data de producción
    # sé que hay artículos mal asignados en categ_id; quizá en otro campo tambien
    
    # al menos en demo company_id está vacío; revisar en producción
    # ,['company_id','!=',False] filtro
    # (28052024) en producción tampoco tiene valor company_id; problema: qué artículos pertenecen y se mueven sólo en la financiera?
    ordersLines = models.execute_kw(db, uid, password,'product.product', 'search_read',
                                 [[['active','=',True],['categ_id','in',['Principal','Servicios y Software']],['type','!=','consu']]],
                                 {'fields':['id','categ_id','code','name','product_tmpl_id','price','price_extra','lst_price','standard_price',
                                            '__last_update','stock_quant_ids','stock_move_ids','qty_available','virtual_available','free_qty',
                                            'incoming_qty','outgoing_qty','orderpoint_ids','bom_line_ids','bom_count','used_in_bom_count',
                                            'mrp_product_qty','sale_avg_price','purchase_avg_price','sale_num_invoiced','purchase_num_invoiced',
                                            'sales_gap','purchase_gap','turnover','total_cost','sale_expected','normal_cost','total_margin',
                                            'expected_margin','total_margin_rate','expected_margin_rate','purchased_product_qty','value_svl',
                                            'quantity_svl','purchase_order_line_ids','sales_count','x_studio_nombre_producto_mix','currency_id',
                                            'cost_currency_id','list_price','uom_id','seller_ids','variant_seller_ids','product_variant_ids',
                                            'product_variant_id','taxes_id','supplier_taxes_id','x_studio_upc']})
    '''
    linea = 0
    for orderLine in ordersLines:
        print(f"item: {linea}: ")
        print(orderLine)
        print('\n')
        linea +=1        
    '''
    campos = []
    for campo in ordersLines[0].keys():
        campos.append(campo)
        
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'product.product'
    
    tmpRow = 1
    tmpColumn = 1
    for _ in range(0, len(campos)):
        ws.cell(row=1, column=_+1).value = campos[_]
        
    tmpRow = tmpRow + 1 #to make content start filling at row 2
    tmpColumn = 1
    
    for order in ordersLines:
        for campo in campos:
            ws.cell(row=tmpRow, column=tmpColumn).value= str(order[campo])
            tmpColumn = tmpColumn + 1
        tmpColumn=1
        tmpRow = tmpRow + 1
    
    wb.save(r"C:\\Users\\ESCH\Desktop\\odoo-852\\excel\\productProduct.xlsx")        
    
    print(f"\n********************\nSe acab´lo que se daba")
    
    
if __name__ == '__main__':
    main()