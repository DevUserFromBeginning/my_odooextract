import xmlrpc.client
import openpyxl
import pprint


'''
Realizado por: Ernesto Caraballo
Fecha actual: 10/05/2024

.- Conectarme a la api (OK)
.- Extraer los registros del modelo: crm.leads
.- Determinas los campos útiles del modelo para cada caso (852, GPOS, etc)
.- Llevar los registros a excel

'''

def main():
    url = 'https://importvzla-import-import-prueba-12847814.dev.odoo.com'
    db = 'importvzla-import-import-prueba-12847814'
    username = 'admin'
    password = 'Import2023!'

    common = xmlrpc.client.ServerProxy('{}/xmlrpc/2/common'.format(url))
    version = common.version()

    uid = common.authenticate(db, username, password, {})

    models = xmlrpc.client.ServerProxy('{}/xmlrpc/2/object'.format(url))
    
    #['company_id', '=', 'Import Import'] # filtro de compañía...aparentemente no funciona
    ordersLines = models.execute_kw(db, uid, password,'stock.picking', 'search_read',
                                 [[['company_id','=','Import Import'],['state','in',['waiting','confirmed','assigned']],['location_dest_id','=','Partner Locations/Customers']]],
                                 {'fields':['id','name','origin','note','backorder_id','backorder_ids','state','group_id','scheduled_date','date_deadline',
                                            'date','date_done','location_id','location_dest_id','move_lines','move_ids_without_package','picking_type_id',
                                            'picking_type_code','use_existing_lots','partner_id','company_id','move_line_ids','move_line_ids_without_package',
                                            'move_line_nosuggest_ids','move_line_exist','show_check_availability','show_validate','product_id',
                                            'show_operations','show_reserved','show_lots_text','__last_update','display_name','create_uid','create_date',
                                            'write_uid','write_date','country_code','purchase_id','sale_id','carrier_price','destination_country_code',
                                            'x_numero','x_studio_numero_de_factura']})
    '''
    for orderLine in ordersLines:
        print(orderLine)
        print('\n')
    '''
    
    campos = []
    for campo in ordersLines[0].keys():
        campos.append(campo)
        
    ###print(campos)
    '''
    temp = 1
    for tmpOrder in ordersLines:
        print(f'\n\n{temp}')
        print(tmpOrder)
        temp += 1
        
    '''
    wb = openpyxl.Workbook()
    wb.create_sheet("stock.picking")
    ws = wb['stock.picking']
    
    tmpRow = 1
    tmpColumn = 1
    for _ in range(0, len(campos)):
        ws.cell(row=1, column=_+1).value = campos[_]
        
    tmpRow = tmpRow + 1 #to make content start filling at row 2
    tmpColumn = 1
    
    for order in ordersLines:
        for campo in campos:
            ws.cell(row=tmpRow, column=tmpColumn).value= str(order[campo])
            tmpColumn = tmpColumn + 1
        tmpColumn=1
        tmpRow = tmpRow + 1
    
    wb.save(r"C:\\Users\\ESCH\Desktop\\odoo_api\\excel\\stockPicking.xlsx")        
    
    print(f"\n********************\nSe acab´lo que se daba")
    
    
if __name__ == '__main__':
    main()