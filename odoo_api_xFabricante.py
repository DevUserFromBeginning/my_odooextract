import xmlrpc.client
import openpyxl


'''
Realizado por: Ernesto Caraballo
Fecha actual: 10/05/2024

.- Conectarme a la api (OK)
.- Extraer los registros del modelo: x_fabricante
.- Determinas los campos útiles del modelo para cada caso (852, GPOS, etc)
.- Llevar los registros a excel

'''

def main():
    url = 'https://importvzla-import-import-prueba-12847814.dev.odoo.com'
    db = 'importvzla-import-import-prueba-12847814'
    username = 'admin'
    password = 'Import2023!'

    common = xmlrpc.client.ServerProxy('{}/xmlrpc/2/common'.format(url))
    version = common.version()

    uid = common.authenticate(db, username, password, {})

    models = xmlrpc.client.ServerProxy('{}/xmlrpc/2/object'.format(url))
    
    #['company_id', '=', 'Import Import'] # filtro de compañía...aparentemente no funciona
    ordersLines = models.execute_kw(db, uid, password,'x_fabricante', 'search_read',
                                 [[]],
                                 {'fields':['id','display_name','x_studio_descripcin_fabricante','x_name','create_date','__last_update']})
    
    '''
    for orderLine in ordersLines:
        print(orderLine)
        print('\n')
    '''
    
    campos = []
    for campo in ordersLines[0].keys():
        campos.append(campo)
        
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "x_fabricante"
    
    tmpRow = 1
    tmpColumn = 1
    for _ in range(0, len(campos)):
        ws.cell(row=1, column=_+1).value = campos[_]
        
    tmpRow = tmpRow + 1 #to make content start filling at row 2
    tmpColumn = 1
    
    for order in ordersLines:
        for campo in campos:
            ws.cell(row=tmpRow, column=tmpColumn).value= str(order[campo])
            tmpColumn = tmpColumn + 1
        tmpColumn=1
        tmpRow = tmpRow + 1
    
    wb.save(r"C:\\Users\\ESCH\Desktop\\odoo\\excel\\x_fabricante.xlsx")        
    
    print(f"\n********************\nSe acab´lo que se daba")
    
    
if __name__ == '__main__':
    main()