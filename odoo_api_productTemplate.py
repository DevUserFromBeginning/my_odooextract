import xmlrpc.client
import openpyxl


'''
Realizado por: Ernesto Caraballo
Fecha actual: 10/05/2024

.- Conectarme a la api (OK)
.- Extraer los registros del modelo: product.template
.- Determinas los campos útiles del modelo para cada caso (852, GPOS, etc)
.- Llevar los registros a excel

'''
'''url = 'https://importvzla-import-import-prueba-12847814.dev.odoo.com'
    db = 'importvzla-import-import-prueba-12847814'
    username = 'admin'
    password = 'Import2023!'''
    
def main():
    url = 'https://importvzla-import-import-prueba-12847814.dev.odoo.com'
    db = 'importvzla-import-import-prueba-12847814'
    username = 'francisco.tellez@import-import.com'
    password = '1129734'

    common = xmlrpc.client.ServerProxy('{}/xmlrpc/2/common'.format(url))
    version = common.version()

    uid = common.authenticate(db, username, password, {})

    models = xmlrpc.client.ServerProxy('{}/xmlrpc/2/object'.format(url))
    
    
    ordersLines = models.execute_kw(db, uid, password,'product.template', 'search_read',
                                 [[['categ_id','in',['Principal','Servicios y Software']],['type','!=','consu'],['name','=like','Hamburguesa de queso']]],
                                 {'fields':['id','default_code','name','categ_id','price','list_price','standard_price',
                                            'x_studio_many2one_field_a487A','x_studio_sublinea_1','x_studio_upc','type','currency_id',
                                            'cost_currency_id','uom_name','seller_ids','product_variant_ids','product_variant_id',
                                            'product_variant_count','__last_update','taxes_id','supplier_taxes_id',
                                            'property_stock_inventory','qty_available','virtual_available','incoming_qty','outgoing_qty',
                                            'nbr_moves_in','nbr_moves_out','nbr_reordering_rules','reordering_min_qty','reordering_max_qty',
                                            'bom_line_ids','bom_ids','bom_count','used_in_bom_count','mrp_product_qty','purchased_product_qty',
                                            'sales_count','x_studio_many2one_field_2qA7w']})
    
    '''
    for orderLine in ordersLines:
        print(orderLine)
        print('\n')
    '''
    campos = []
    for campo in ordersLines[0].keys():
        campos.append(campo)
        
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "product.template"
    
    tmpRow = 1
    tmpColumn = 1
    for _ in range(0, len(campos)):
        ws.cell(row=1, column=_+1).value = campos[_]
        
    tmpRow = tmpRow + 1 #to make content start filling at row 2
    tmpColumn = 1
    
    for order in ordersLines:
        for campo in campos:
            ws.cell(row=tmpRow, column=tmpColumn).value= str(order[campo])
            tmpColumn = tmpColumn + 1
        tmpColumn=1
        tmpRow = tmpRow + 1
    
    wb.save(r"C:\\Users\\ESCH\Desktop\\odoo-852\\excel\\productTemplate.xlsx")        
    
    print(f"\n********************\nSe acab´lo que se daba")
    
    
if __name__ == '__main__':
    main()